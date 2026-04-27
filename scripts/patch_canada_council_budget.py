#!/usr/bin/env python3
"""
Patch Canada Council budget workbook in place.
Only updates content for the specified fixes; does not rebuild or change formatting.
"""
from pathlib import Path
from openpyxl import load_workbook

VAULT_BASE = Path.home() / "Library/Mobile Documents/iCloud~md~obsidian/Documents"
OUTPUT_FILE = VAULT_BASE / "Fundraising/Canada Council/Canada Council - Two-Grant Budget.xlsx"


def patch_equipment_note(ws):
    """AC_Expenses: Clarify equipment as multiple items each under $2,000 (Option A)."""
    ws["D14"].value = (
        "All $2,500 from Canada Council; no applicant match. "
        "Split as: SSDs ($1,200) + card readers ($800) + mounts ($500). "
        "Under 15% of request and under $2,000 per item per Council guidelines."
    )


def patch_budget_checks_equipment(ws):
    """Budget_Checks: Use WARNING instead of FAIL for equipment >$2,000 and add tip."""
    ws["B3"].value = '=IF(AC_Expenses!C14<=2000,"OK","WARNING")'
    ws["C3"].value = (
        '=IF(AC_Expenses!C14>2000,"Equipment $2,500: document per-item breakdown in AC_Expenses note (e.g. SSDs $1,200, readers $800, mounts $500). Each item ≤$2,000.","")'
    )


def patch_inkind_mapping(ws):
    """InKind_Mapping: Replace with one clean version (no duplicate/corrupted rows)."""
    for row in range(1, 25):
        for col in range(1, 4):
            try:
                ws.cell(row, col).value = None
            except Exception:
                pass
    content = [
        ("ARTISTIC CREATION IN-KIND SUPPORT",),
        (),
        ("In-kind contribution from partners (if available):",),
        ("Disability Org A: Studio space access (~$2,000 value)",),
        ("Community Center E: Participant refreshments & hospitality (~$1,500 value)",),
        ("Video Professional C: Technical consultation (~$500 value)",),
        (),
        ("Total in-kind: ~$4,000 (supplementary; not substituting Canada Council request)",),
        (),
        ("SECTOR SUPPORT PHASE",),
        ("In-kind: meeting space, staff time from partner orgs (add if confirmed).",),
    ]
    for i, row in enumerate(content, 1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j).value = val


def main():
    if not OUTPUT_FILE.exists():
        print(f"File not found: {OUTPUT_FILE}")
        return
    wb = load_workbook(OUTPUT_FILE)
    patch_equipment_note(wb["AC_Expenses"])
    patch_budget_checks_equipment(wb["Budget_Checks"])
    patch_inkind_mapping(wb["InKind_Mapping"])
    wb.save(OUTPUT_FILE)
    print(f"Patched: {OUTPUT_FILE}")
    print("  - AC_Expenses D14: Equipment note (Option A breakdown)")
    print("  - Budget_Checks B3/C3: Equipment check WARNING + tip")
    print("  - InKind_Mapping: Clean single version")


if __name__ == "__main__":
    main()
