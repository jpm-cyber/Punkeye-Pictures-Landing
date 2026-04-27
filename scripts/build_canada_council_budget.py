#!/usr/bin/env python3
"""
Build Canada Council Two-Grant Budget workbook.
Saves to Obsidian vault: Fundraising/Canada Council/Canada Council - Two-Grant Budget.xlsx

All subtotal/total and Budget_Checks cells use formulas. If they appear blank when you open
the file: open in Microsoft Excel and use Formulas > Calculate Now (or Ctrl+Alt+F9 / Cmd+Option+F9).
The workbook is set to fullCalcOnLoad so Excel should recalculate on open.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.reader.excel import load_workbook

# Output path (Obsidian vault)
VAULT_BASE = Path.home() / "Library/Mobile Documents/iCloud~md~obsidian/Documents"
OUTPUT_DIR = VAULT_BASE / "Fundraising/Canada Council"
OUTPUT_FILE = OUTPUT_DIR / "Canada Council - Two-Grant Budget.xlsx"

AC_REQUEST = 75000
SS_REQUEST = 100000
MAX_NOTE_LEN = 250
EQUIPMENT_PCT_MAX = 0.15
CONTINGENCY_PCT_MAX = 0.10
EQUIPMENT_PER_ITEM_MAX = 2000


def notes_under_250():
    """All budget notes trimmed to under 250 characters."""
    return {
        "ac_lead_artist": "Lead artist fee for ~54 days over 6 months (artistic research, prompts/constraints, weekly dailies review, editing direction) at blended rate aligned with CARFAC-RAAV and Vancouver media arts practice.",
        "ac_participants": "30 neurodivergent participants x $600 each as artist honoraria for sustained creative labour over 6 months. Co-creators retain authorship.",
        "ac_editor": "Editor fee for 3-5 short process films (3-8 min each) from participant media, aligned with mid-range Vancouver post-production.",
        "ac_honoraria_elders": "Not applicable to this project.",
        "ac_materials": "Backup storage and consumables for 6 months of dailies (drives, SD cards, cables) plus participant-facing guide printing.",
        "ac_equipment": "All $2,500 sought from Canada Council; no applicant match. Capture/storage (SSDs, card readers, mounts). Under 15% of request and under $2,000 per item per Council guidelines.",
        "ac_contingency": "Contingency for 6-month phase: replacement devices, additional editing days, unexpected technical support at partner sites. Capped at under 10% of total (8.7%).",
        "ac_travel": "Local Lower Mainland travel for lead artist and editor to partner programs and screenings; transit, mileage, accessible taxi.",
        "ac_accessibility": "Captioning, relaxed performance, access coordination for community screenings (support workers/interpretation where not from partners).",
        "ac_revenue_main": "Primary and only cash revenue source for this 6-month artistic research-creation project. No other public, private, or earned revenues confirmed or anticipated.",
        "ss_artistic_note": "Artistic labour in Artistic Creation grant only; this file = systems, PM, sector replication.",
        "ss_pro_fees": "Systems/PM ~20 hrs/week over 12 months (~1,040 hrs) at blended rate; plus contracted UX/privacy advisory. Workflow design, documentation, multi-site coordination. No artistic creation.",
        "ss_documentation": "Turnkey toolkit: facilitator modules, checklists, onboarding scripts, workflow diagrams, web resources.",
        "ss_materials": "Software: annual subscriptions for Punkeye to produce/host toolkit (diagramming, collaborative platforms, hosting), no single tool over $500/yr. Not capital equipment; aligned with Council policy.",
        "ss_venue": "Rental of meeting and training spaces (or equivalent online platforms) for sector workshops and implementation labs.",
        "ss_travel": "Local and regional travel to disability-support and arts organizations for implementation meetings and workshops over 12 months.",
        "ss_per_diem": "Per diems and occasional accommodation for longer trips where in-person sector visits are needed.",
        "ss_admin": "Admin/comms: outreach to organizations, scheduling pilot implementations and workshops, managing registrations, preparing sector-facing reports and updates. Contract or part-time role.",
        "ss_safety": "Supplementary insurance and risk management advice for data privacy, storage, and participant safety during multi-site workflow implementation.",
        "ss_childcare": "Childcare/dependent care for lead systems developer; essential for documentation and sector outreach delivery over 12 months. Eligible Council expense.",
        "ss_accessibility": "Accessibility for sector-facing workshops and materials: interpretation, captioning, and accessible formats so organizations can fully participate.",
        "ss_revenue_main": "Primary and only cash revenue source for this 12-month systems and capacity-building project; no other public, private, or earned revenues confirmed.",
    }


def ensure_notes_under_250(notes_dict):
    for k, v in notes_dict.items():
        if len(v) > MAX_NOTE_LEN:
            notes_dict[k] = v[: MAX_NOTE_LEN - 3] + "..."
    return notes_dict


def build_cover(ws):
    ws["A1"] = "Punkeye Pictures: Two-Grant Strategy & Detailed Budgets"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Artistic Creation + Sector Support"
    ws["A3"] = "Prepared for: Joe Moulins / Punkeye Pictures"
    ws["A4"] = "Date: March 5, 2026"
    ws["A5"] = "Total Requested Across Both Grants: $175,000 ($75,000 + $100,000)"
    ws["A7"] = "Project overview"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = (
        "Disability-led media arts initiative with two dimensions: "
        "(1) Artistic Inquiry & Creation — 6-month research-creation pilot with 25-30 participants, "
        "process films and community screenings; (2) Sector Innovation & Replication — 12-month phase "
        "documenting and sharing the workflow as a turnkey toolkit for disability-support and arts organizations."
    )
    ws["A8"].alignment = Alignment(wrap_text=True)
    ws["A10"] = "Grants"
    ws["A10"].font = Font(bold=True)
    ws["A11"] = "Grant"
    ws["B11"] = "Program"
    ws["C11"] = "Duration"
    ws["D11"] = "Request"
    ws["A12"] = "Artistic Creation"
    ws["B12"] = "Explore & Create"
    ws["C12"] = "6 months"
    ws["D12"] = 75000
    ws["A13"] = "Sector Support"
    ws["B13"] = "Innovation and Development"
    ws["C13"] = "12 months"
    ws["D13"] = 100000
    ws["A15"] = "Expense delineation"
    ws["A15"].font = Font(bold=True)
    ws["A16"] = "Artistic Creation ONLY pays for: Lead artist fee, participant honoraria, editor, materials/equipment for capture and storage, travel to partner sites and screenings, accessibility for community screenings, contingency. Does NOT pay for: system architecture, sector documentation, sector workshops."
    ws["A16"].alignment = Alignment(wrap_text=True)
    ws["A17"] = "Sector Support ONLY pays for: Systems/PM fee, advisers, documentation/toolkit, venue for sector workshops, admin/comms, travel for sector outreach, safety/insurance, childcare, accessibility for sector events. Does NOT pay for: artistic research, editing, participant sessions, community screenings."
    ws["A17"].alignment = Alignment(wrap_text=True)
    ws["A18"] = "Equipment (Artistic Creation): All equipment amount requested is from Canada Council; no applicant/in-kind match required for that line."
    ws["A18"].alignment = Alignment(wrap_text=True)
    ws["A19"] = "Before submission: Confirm Sector Support program ceiling ($100K vs $50K typical) with program officer. Add in-kind support if partners provide space, advisory time, or hosting."
    ws["A19"].alignment = Alignment(wrap_text=True)
    ws["A19"].font = Font(italic=True)
    ws["A20"] = "Budget tabs: Amounts are in column C (AC_Expenses, SS_Expenses); column B (AC_Income, SS_Income). All subtotals and totals use formulas. If cells appear blank, open in Excel and use Formulas > Calculate Now (or Ctrl+Alt+F9)."
    ws["A20"].alignment = Alignment(wrap_text=True)
    ws["A20"].font = Font(italic=True)
    ws.column_dimensions["A"].width = 100


def build_ac_expenses(ws, notes):
    ws["A1"] = "Category"
    ws["B1"] = "Portal line / Line item"
    ws["C1"] = "Amount"
    ws["D1"] = "Note (max 250 chars)"
    for c in range(1, 5):
        ws.cell(1, c).font = Font(bold=True)
    # Artistic and project: rows 2-8, subtotal row 9
    rows = [
        (2, "Artistic and project", "Artists and professional fees", 27000, notes["ac_lead_artist"]),
        (3, None, "Artistic salaries (participants)", 18000, notes["ac_participants"]),
        (4, None, "Other artistic and project expenses A (editor)", 14000, notes["ac_editor"]),
        (5, None, "Honoraria for Elders and Knowledge Keepers", 0, notes["ac_honoraria_elders"]),
        (6, None, "Copyright, reproduction and royalties", 0, None),
        (7, None, "Other artistic and project expenses B", 0, None),
        (8, None, "Other artistic and project expenses C", 0, None),
    ]
    for r, cat, line, amt, note in rows:
        if cat:
            ws.cell(r, 1).value = cat
        ws.cell(r, 2).value = line
        ws.cell(r, 3).value = amt
        if note:
            ws.cell(r, 4).value = note
    ws["A9"] = "Subtotal Artistic and project"
    ws["C9"] = "=SUM(C2:C8)"
    ws["C9"].number_format = '"$"#,##0'
    # Production, technical, programming: rows 11-20, subtotal row 21. Equipment row 14, contingency row 18.
    prod_rows = [
        (11, "Production, technical and programming", "Production, technical and programming professional fees", 0, None),
        (12, None, "Documentation", 0, None),
        (13, None, "Materials and supplies", 2000, notes["ac_materials"]),
        (14, None, "Equipment purchase", 2500, notes["ac_equipment"]),
        (15, None, "Equipment rental", 0, None),
        (16, None, "Venue, studio, facility rental", 0, None),
        (17, None, "Other production, technical and programming expenses A (contingency)", 6500, notes["ac_contingency"]),
        (18, None, "Registration fees for professional development", 0, None),
        (19, None, "Other production B", 0, None),
        (20, None, "Other production C", 0, None),
    ]
    for r, cat, line, amt, note in prod_rows:
        if cat:
            ws.cell(r, 1).value = cat
        ws.cell(r, 2).value = line
        ws.cell(r, 3).value = amt
        if note:
            ws.cell(r, 4).value = note
    ws["A21"] = "Subtotal Production, technical, programming"
    ws["C21"] = "=SUM(C11:C20)"
    ws["C21"].number_format = '"$"#,##0'
    # Travel: rows 23-24, subtotal row 25
    ws.cell(23, 1).value = "Travel"
    ws.cell(23, 2).value = "Personnel Travel"
    ws.cell(23, 3).value = 2500
    ws.cell(23, 4).value = notes["ac_travel"]
    ws.cell(24, 2).value = "Per diem and accommodations"
    ws.cell(24, 3).value = 0
    ws["A25"] = "Subtotal Travel"
    ws["C25"] = "=SUM(C23:C24)"
    ws["C25"].number_format = '"$"#,##0'
    # Admin: row 26
    ws.cell(26, 1).value = "Administrative, marketing and communications"
    ws.cell(26, 3).value = 0
    # Other: row 27
    ws.cell(27, 1).value = "Other expenses"
    ws.cell(27, 3).value = 0
    # Accessibility: row 28-29, subtotal row 30
    ws.cell(28, 1).value = "Accessibility"
    ws.cell(28, 2).value = "Public Accessibility costs"
    ws.cell(28, 3).value = 2500
    ws.cell(28, 4).value = notes["ac_accessibility"]
    ws["A30"] = "Subtotal Accessibility"
    ws["C30"] = "=C28"
    ws["C30"].number_format = '"$"#,##0'
    # Total: row 32
    ws["A32"] = "TOTAL EXPENSES"
    ws["A32"].font = Font(bold=True)
    ws["C32"] = "=SUM(C9,C21,C25,C26,C27,C30)"
    ws["C32"].number_format = '"$"#,##0'
    for r in (9, 21, 25, 26, 27, 30):
        ws.cell(r, 3).number_format = '"$"#,##0'
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    return {"equipment_row": 14, "contingency_row": 17, "total_row": 32}


def build_ac_income(ws, notes):
    ws["A1"] = "Line"
    ws["B1"] = "Amount"
    ws["C1"] = "Note (max 250 chars)"
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True)
    ws["A2"] = "Canada Council for the Arts (this application)"
    ws["B2"] = 75000
    ws["C2"] = notes["ac_revenue_main"]
    ws["A3"] = "Other Canada Council"
    ws["B3"] = 0
    ws["A4"] = "Other federal, provincial, municipal"
    ws["B4"] = 0
    ws["A5"] = "Earned revenue"
    ws["B5"] = 0
    ws["A6"] = "Private sector"
    ws["B6"] = 0
    ws["A7"] = "In-kind support"
    ws["B7"] = 0
    ws["A8"] = "Applicant contribution"
    ws["B8"] = 0
    ws["A9"] = "TOTAL REVENUE"
    ws["A9"].font = Font(bold=True)
    ws["B9"] = "=SUM(B2:B8)"
    ws["B9"].number_format = '"$"#,##0'
    ws["A10"] = "NET (Revenue - Expenses)"
    ws["A10"].font = Font(bold=True)
    ws["B10"] = "=B9-'AC_Expenses'!C32"
    ws["B10"].number_format = '"$"#,##0'


def build_ss_expenses(ws, notes):
    ws["A1"] = "Category"
    ws["B1"] = "Portal line / Line item"
    ws["C1"] = "Amount"
    ws["D1"] = "Note (max 250 chars)"
    for c in range(1, 5):
        ws.cell(1, c).font = Font(bold=True)
    # Artistic and project: row 2 only (clarification line), subtotal row 3
    ws["A2"] = "Artistic and project"
    ws["B2"] = "All artistic labour in Artistic Creation grant"
    ws["C2"] = 0
    ws["D2"] = notes["ss_artistic_note"]
    ws["A3"] = "Subtotal Artistic and project"
    ws["C3"] = "=SUM(C2:C2)"
    ws["C3"].number_format = '"$"#,##0'
    # Production, technical and programming: rows 5-8, subtotal row 9
    ws["A4"] = "Production, technical and programming"
    ws["B5"] = "Production, technical and programming professional fees"
    ws["C5"] = 60000
    ws["D5"] = notes["ss_pro_fees"]
    ws["B6"] = "Documentation"
    ws["C6"] = 10000
    ws["D6"] = notes["ss_documentation"]
    ws["B7"] = "Materials and supplies"
    ws["C7"] = 3000
    ws["D7"] = notes["ss_materials"]
    ws["B8"] = "Venue, studio, facility rental"
    ws["C8"] = 5000
    ws["D8"] = notes["ss_venue"]
    ws["A9"] = "Subtotal Production, technical, programming"
    ws["C9"] = "=SUM(C5:C8)"
    ws["C9"].number_format = '"$"#,##0'
    # Travel: rows 11-12, subtotal row 13
    ws["A10"] = "Travel"
    ws["B11"] = "Personnel Travel"
    ws["C11"] = 3000
    ws["D11"] = notes["ss_travel"]
    ws["B12"] = "Per diem and accommodations"
    ws["C12"] = 1000
    ws["D12"] = notes["ss_per_diem"]
    ws["A13"] = "Subtotal Travel"
    ws["C13"] = "=SUM(C11:C12)"
    ws["C13"].number_format = '"$"#,##0'
    # Administrative: row 15
    ws["A14"] = "Administrative, marketing and communications"
    ws["B15"] = "Administrative, marketing and communications professional fees"
    ws["C15"] = 8000
    ws["D15"] = notes["ss_admin"]
    ws["A16"] = "Subtotal Administrative"
    ws["C16"] = "=C15"
    ws["C16"].number_format = '"$"#,##0'
    # Other: rows 18-19
    ws["A17"] = "Other expenses"
    ws["B18"] = "Safety-related costs"
    ws["C18"] = 2000
    ws["D18"] = notes["ss_safety"]
    ws["B19"] = "Childcare/dependent care costs"
    ws["C19"] = 3000
    ws["D19"] = notes["ss_childcare"]
    ws["A20"] = "Subtotal Other"
    ws["C20"] = "=SUM(C18:C19)"
    ws["C20"].number_format = '"$"#,##0'
    # Accessibility: row 22, subtotal row 23
    ws["A21"] = "Accessibility"
    ws["B22"] = "Public Accessibility costs"
    ws["C22"] = 5000
    ws["D22"] = notes["ss_accessibility"]
    ws["A23"] = "Subtotal Accessibility"
    ws["C23"] = "=C22"
    ws["C23"].number_format = '"$"#,##0'
    # TOTAL: row 25
    ws["A25"] = "TOTAL EXPENSES"
    ws["A25"].font = Font(bold=True)
    ws["C25"] = "=SUM(C3,C9,C13,C16,C20,C23)"
    ws["C25"].number_format = '"$"#,##0'
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 60
    return {"total_row": 25}


def build_ss_income(ws, notes):
    ws["A1"] = "Line"
    ws["B1"] = "Amount"
    ws["C1"] = "Note (max 250 chars)"
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True)
    ws["A2"] = "Canada Council for the Arts (this application)"
    ws["B2"] = 100000
    ws["C2"] = notes["ss_revenue_main"]
    ws["A3"] = "Other Canada Council"
    ws["B3"] = 0
    ws["A4"] = "Other federal, provincial, municipal"
    ws["B4"] = 0
    ws["A5"] = "Earned revenue"
    ws["B5"] = 0
    ws["A6"] = "Private sector"
    ws["B6"] = 0
    ws["A7"] = "In-kind support"
    ws["B7"] = 0
    ws["A8"] = "Applicant contribution"
    ws["B8"] = 0
    ws["A9"] = "TOTAL REVENUE"
    ws["A9"].font = Font(bold=True)
    ws["B9"] = "=SUM(B2:B8)"
    ws["B9"].number_format = '"$"#,##0'
    ws["A10"] = "NET (Revenue - Expenses)"
    ws["A10"].font = Font(bold=True)
    ws["B10"] = "=B9-'SS_Expenses'!C25"
    ws["B10"].number_format = '"$"#,##0'


def build_budget_checks(ws):
    ws["A1"] = "Check"
    ws["B1"] = "Result"
    ws["C1"] = "Detail / Tip"
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True)
    # AC equipment % <= 15% (equipment cell C14)
    ws["A2"] = "AC: Equipment total <= 15% of request"
    ws["B2"] = '=IF(AC_Expenses!C14/75000<=0.15,"OK","FAIL")'
    ws["C2"] = '=IF(AC_Expenses!C14/75000>0.15,"Reduce equipment or request; keep total <= 15% (max $11,250).","")'
    # AC equipment per item <= 2000
    ws["A3"] = "AC: No single equipment item > $2,000"
    ws["B3"] = '=IF(AC_Expenses!C14<=2000,"OK","FAIL")'
    ws["C3"] = '=IF(AC_Expenses!C14>2000,"Each equipment item must be <= $2,000 per Council guidelines.","")'
    # AC contingency <= 10% (contingency cell C17)
    ws["A4"] = "AC: Contingency <= 10% of total"
    ws["B4"] = '=IF(AC_Expenses!C17/75000<=0.1,"OK","FAIL")'
    ws["C4"] = '=IF(AC_Expenses!C17/75000>0.1,"Contingency must be under 10% of request (max $7,500).","")'
    # AC total = 75000 (total cell C32)
    ws["A5"] = "AC: Total expenses = $75,000"
    ws["B5"] = '=IF(ABS(AC_Expenses!C32-75000)<1,"OK","FAIL")'
    ws["C5"] = '=IF(ABS(AC_Expenses!C32-75000)>=1,"Adjust line items so total equals $75,000.","")'
    # AC revenue = 75000
    ws["A6"] = "AC: Total revenue = $75,000"
    ws["B6"] = '=IF(ABS(AC_Income!B9-75000)<1,"OK","FAIL")'
    ws["C6"] = '=IF(ABS(AC_Income!B9-75000)>=1,"Revenue must equal request ($75,000).","")'
    # SS total = 100000 (SS total in C25)
    ws["A7"] = "SS: Total expenses = $100,000"
    ws["B7"] = '=IF(ABS(SS_Expenses!C25-100000)<1,"OK","FAIL")'
    ws["C7"] = '=IF(ABS(SS_Expenses!C25-100000)>=1,"Adjust line items so total equals $100,000.","")'
    # SS revenue = 100000
    ws["A8"] = "SS: Total revenue = $100,000"
    ws["B8"] = '=IF(ABS(SS_Income!B9-100000)<1,"OK","FAIL")'
    ws["C8"] = '=IF(ABS(SS_Income!B9-100000)>=1,"Revenue must equal request ($100,000).","")'
    # Note length checks
    ws["A9"] = "AC: Lead artist note <= 250 chars"
    ws["B9"] = '=IF(LEN(AC_Expenses!D2)<=250,"OK","WARNING")'
    ws["C9"] = '=IF(LEN(AC_Expenses!D2)>250,"Shorten note for portal (max 250 characters).","")'
    ws["A10"] = "SS: Pro fees note <= 250 chars"
    ws["B10"] = '=IF(LEN(SS_Expenses!D5)<=250,"OK","WARNING")'
    ws["C10"] = '=IF(LEN(SS_Expenses!D5)>250,"Shorten note for portal (max 250 characters).","")'
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 55


def add_data_validation(wb):
    # Per-sheet validations (each sheet gets its own DataValidation with correct sqref)
    # AC_Expenses: amounts C2:C32, notes D2:D28
    ac_exp = wb["AC_Expenses"]
    ac_exp.add_data_validation(
        DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showInputMessage=True,
            promptTitle="Amount",
            prompt="Canada Council: amounts must be >= 0.",
            sqref="C2:C32",
        )
    )
    ac_exp.add_data_validation(
        DataValidation(
            type="textLength",
            operator="lessThanOrEqual",
            formula1="250",
            allow_blank=True,
            showInputMessage=True,
            promptTitle="Note length",
            prompt="Portal notes must be 250 characters or fewer.",
            sqref="D2:D28",
        )
    )
    ac_exp.add_data_validation(
        DataValidation(
            type="decimal",
            operator="between",
            formula1="0",
            formula2="11250",
            allow_blank=True,
            showInputMessage=True,
            promptTitle="Equipment (Canada Council)",
            prompt="Equipment total <= 15% of request ($11,250); each item <= $2,000.",
            sqref="C14",
        )
    )
    # AC_Income: amounts B2:B8, notes C2:C8
    ac_inc = wb["AC_Income"]
    ac_inc.add_data_validation(
        DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showInputMessage=True, prompt="Amount >= 0.", sqref="B2:B8")
    )
    ac_inc.add_data_validation(
        DataValidation(type="textLength", operator="lessThanOrEqual", formula1="250", allow_blank=True, showInputMessage=True, prompt="Note <= 250 chars.", sqref="C2:C8")
    )
    # SS_Expenses: amounts C2:C28, notes D2:D26
    ss_exp = wb["SS_Expenses"]
    ss_exp.add_data_validation(
        DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showInputMessage=True, prompt="Amount >= 0.", sqref="C2:C25")
    )
    ss_exp.add_data_validation(
        DataValidation(type="textLength", operator="lessThanOrEqual", formula1="250", allow_blank=True, showInputMessage=True, prompt="Note <= 250 chars.", sqref="D2:D22")
    )
    # SS_Income: amounts B2:B8, notes C2:C8
    ss_inc = wb["SS_Income"]
    ss_inc.add_data_validation(
        DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showInputMessage=True, prompt="Amount >= 0.", sqref="B2:B8")
    )
    ss_inc.add_data_validation(
        DataValidation(type="textLength", operator="lessThanOrEqual", formula1="250", allow_blank=True, showInputMessage=True, prompt="Note <= 250 chars.", sqref="C2:C8")
    )


def add_conditional_formatting(wb):
    light_red = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    # AC_Expenses: equipment C14 > 11250 (15%) or > 2000 per item
    ac = wb["AC_Expenses"]
    ac.conditional_formatting.add(
        "C14",
        CellIsRule(operator="greaterThan", formula=["11250"], fill=light_red),
    )
    ac.conditional_formatting.add(
        "C14",
        CellIsRule(operator="greaterThan", formula=["2000"], fill=light_red),
    )
    # AC_Expenses: contingency C17 > 7500 (10% of 75000)
    ac.conditional_formatting.add(
        "C17",
        CellIsRule(operator="greaterThan", formula=["7500"], fill=light_red),
    )
    # AC_Expenses: total C32 not 75000
    ac.conditional_formatting.add(
        "C32",
        FormulaRule(formula=["C32<>75000"], fill=light_red),
    )
    # SS_Expenses: total not 100000 (total in C25)
    ss = wb["SS_Expenses"]
    ss.conditional_formatting.add(
        "C25",
        FormulaRule(formula=["C25<>100000"], fill=light_red),
    )
    # AC_Income: total revenue not 75000
    ac_inc = wb["AC_Income"]
    ac_inc.conditional_formatting.add(
        "B9",
        FormulaRule(formula=["B9<>75000"], fill=light_red),
    )
    # SS_Income: total revenue not 100000
    ss_inc = wb["SS_Income"]
    ss_inc.conditional_formatting.add(
        "B9",
        FormulaRule(formula=["B9<>100000"], fill=light_red),
    )


def build_ac_cashflow(ws):
    """Artistic Creation grant - 6-month cash flow timeline."""
    ws["A1"] = "ARTISTIC CREATION GRANT - Cash Flow Timeline (6 months)"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Category", "Month 1", "Month 2", "Month 3", "Month 4", "Month 5", "Month 6", "TOTAL"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    rows = [
        ("Lead artist", 4500, 4500, 4500, 4500, 4500, 4500, 27000),
        ("Participant honoraria", 3000, 3000, 3000, 3000, 3000, 3000, 18000),
        ("Editor", 0, 0, 7000, 7000, 0, 0, 14000),
        ("Materials/supplies", 500, 300, 300, 300, 300, 200, 2000),
        ("Equipment", 2500, 0, 0, 0, 0, 0, 2500),
        ("Travel", 500, 500, 500, 500, 500, 0, 2500),
        ("Accessibility", 0, 0, 1250, 1250, 0, 0, 2500),
        ("Contingency", 0, 0, 0, 0, 0, 6500, 6500),
    ]
    for i, (cat, m1, m2, m3, m4, m5, m6, tot) in enumerate(rows, 3):
        ws.cell(i, 1, cat)
        ws.cell(i, 2, m1)
        ws.cell(i, 3, m2)
        ws.cell(i, 4, m3)
        ws.cell(i, 5, m4)
        ws.cell(i, 6, m5)
        ws.cell(i, 7, m6)
        ws.cell(i, 8, tot)
    ws.cell(11, 1, "MONTHLY TOTAL")
    ws.cell(11, 1).font = Font(bold=True)
    ws["B11"] = "=SUM(B3:B10)"
    ws["C11"] = "=SUM(C3:C10)"
    ws["D11"] = "=SUM(D3:D10)"
    ws["E11"] = "=SUM(E3:E10)"
    ws["F11"] = "=SUM(F3:F10)"
    ws["G11"] = "=SUM(G3:G10)"
    ws["H11"] = "=SUM(H3:H10)"
    ws.cell(12, 1, "Cumulative")
    ws.cell(12, 1).font = Font(bold=True)
    ws["B12"] = "=B11"
    ws["C12"] = "=B12+C11"
    ws["D12"] = "=C12+D11"
    ws["E12"] = "=D12+E11"
    ws["F12"] = "=E12+F11"
    ws["G12"] = "=F12+G11"
    for col in range(2, 9):
        ws.cell(11, col).number_format = '"$"#,##0'
        ws.cell(12, col).number_format = '"$"#,##0'
    for r in range(3, 11):
        for c in range(2, 9):
            ws.cell(r, c).number_format = '"$"#,##0'
    ws.column_dimensions["A"].width = 22


def build_ss_cashflow(ws):
    """Sector Support grant - 12-month cash flow (simplified distribution)."""
    from openpyxl.utils import get_column_letter
    ws["A1"] = "SECTOR SUPPORT GRANT - Cash Flow Timeline (12 months)"
    ws["A1"].font = Font(bold=True, size=12)
    for col in range(1, 14):
        ws.cell(2, col, f"M{col}" if col <= 12 else "TOTAL").font = Font(bold=True)
    # Rows: Systems/PM 5k/mo, Documentation back-loaded, Materials even, Venue workshops M4-M10, Travel/Admin/Other/Accessibility even
    ws.cell(3, 1, "Systems/PM + advisory")
    for col in range(2, 14):
        ws.cell(3, col, 5000 if col <= 12 else 60000)
    ws.cell(4, 1, "Documentation")
    for col in range(2, 14):
        ws.cell(4, col, 0 if col < 8 else (2500 if col == 9 else 1667 if col in (10, 11, 12) else 0) if col <= 12 else 10000)
    ws.cell(4, 9, 2500)
    ws.cell(4, 10, 2500)
    ws.cell(4, 11, 2500)
    ws.cell(4, 12, 2500)
    ws.cell(4, 13, 10000)
    ws.cell(5, 1, "Materials")
    for col in range(2, 14):
        ws.cell(5, col, 250 if col <= 12 else 3000)
    ws.cell(6, 1, "Venue (workshops)")
    for col in range(2, 14):
        ws.cell(6, col, (500 if 4 <= col <= 11 else 0) if col <= 12 else 5000)
    ws.cell(7, 1, "Travel")
    for col in range(2, 14):
        ws.cell(7, col, (333 if col <= 11 else 337) if col <= 12 else 4000)
    ws.cell(8, 1, "Admin/comms")
    for col in range(2, 14):
        ws.cell(8, col, 667 if col <= 12 else 8000)
    ws.cell(9, 1, "Other (safety + childcare)")
    for col in range(2, 14):
        ws.cell(9, col, 417 if col <= 12 else 5000)
    ws.cell(10, 1, "Accessibility")
    for col in range(2, 14):
        ws.cell(10, col, 417 if col <= 12 else 5000)
    ws.cell(11, 1, "MONTHLY TOTAL")
    ws.cell(11, 1).font = Font(bold=True)
    for col in range(2, 14):
        ws.cell(11, col).value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}10)"
    ws.column_dimensions["A"].width = 24


def build_risk_analysis(ws):
    """Risk register for Artistic Creation grant."""
    ws["A1"] = "RISK REGISTER - Artistic Creation Grant"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["RISK", "PROBABILITY", "IMPACT", "MITIGATION", "CONTINGENCY BUDGET"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    rows = [
        ("Participant drop-off (expected 25-30, actual 15). Budget impact: -$3,000", "Medium", "High", "Clear onboarding, community integration", "If <20 participants, reduce editing scope. Reserve: $2,000-3,000"),
        ("Equipment failure (SSD corruption, loss). Budget impact: -$1,500", "Low", "Medium", "Backup drives, offsite backup (materials)", "Replacement in contingency. Reserve: $1,500"),
        ("Editor unavailable (illness, other contract). Budget impact: -$5,000", "Low", "High", "Contract 2 editors in advance; backup identified", "Emergency post-prod from known vendors. Reserve: $3,000"),
        ("Partner site access blocked (e.g. emergency). Budget impact: -$2,000", "Low", "Medium", "Early confirmation of access dates", "Virtual alternative for screenings. Reserve: $1,000"),
        ("Accessibility costs exceed budget (interpretation, captioning). Budget impact: -$1,000", "Medium", "Low", "Early quotes from access providers", "Additional contingency. Reserve: $1,000"),
    ]
    for i, row in enumerate(rows, 3):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    ws.cell(8, 1, "TOTAL CONTINGENCY RESERVED: $6,500 (8.7% of budget). Individual scenario reserves total $8,500-$9,500; the $6,500 budget means prioritization if multiple issues arise (participant retention and editor continuity prioritized).")
    ws.cell(8, 1).font = Font(bold=True)
    ws.merge_cells("A8:E8")
    for c in range(1, 6):
        ws.column_dimensions[ws.cell(2, c).column_letter].width = 18 if c == 1 else 12 if c <= 3 else 28


def build_inkind_mapping(ws):
    """Funding and in-kind support mapping."""
    ws["A1"] = "FUNDING & IN-KIND SUPPORT MAPPING"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "ARTISTIC CREATION PHASE"
    ws["A2"].font = Font(bold=True)
    ws["A3"] = "Essential items from Canada Council:"
    ws["A4"] = "Lead artist fee: $27,000 | Participant honoraria: $18,000 | Editor: $14,000"
    ws["A5"] = "Equipment: $2,500 | Materials: $2,000 | Travel: $2,500 | Accessibility: $2,500 | Contingency: $6,500"
    ws["A6"] = "Total from Council: $75,000"
    ws["A8"] = "Potential In-Kind (if available):"
    ws["A8"].font = Font(bold=True)
    ws["A9"] = "Item"
    ws["B9"] = "Value Est."
    ws["C9"] = "Partner(s)"
    for c in range(1, 4):
        ws.cell(9, c).font = Font(bold=True)
    ws["A10"] = "Workspace/studio access"
    ws["B10"] = 3000
    ws["C10"] = "[Disability org A] [Arts org B]"
    ws["A11"] = "Technical advisory (pro bono)"
    ws["B11"] = 2000
    ws["C11"] = "[Video professional C]"
    ws["A12"] = "Equipment loan (cameras, etc)"
    ws["B12"] = 1500
    ws["C12"] = "[Media arts group D]"
    ws["A13"] = "Venue for community screening"
    ws["B13"] = 2000
    ws["C13"] = "[Community center E]"
    ws["A14"] = "Potential In-Kind Total: $8,500 (optional; strengthens application)"
    ws["A15"] = "Action: Confirm with partners whether they can contribute in-kind. Add to Income if confirmed."
    ws["A17"] = "SECTOR SUPPORT PHASE"
    ws["A17"].font = Font(bold=True)
    ws["A18"] = "Total from Council: $100,000. In-kind: meeting space, staff time from partner orgs (add if confirmed)."
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 35


def build_expense_justification(ws):
    """Detailed line-by-line justification beyond 250-char notes."""
    ws["A1"] = "EXPENSE JUSTIFICATION - Detailed Breakdown"
    ws["A1"].font = Font(bold=True, size=12)
    blocks = [
        ("LINE: Artists and professional fees - $27,000", [
            "Role: Lead Artist (Research Director + Editing Direction)",
            "Duration: 6 months | Days: ~54 (avg ~2 days/week) | Rate: $500/day blended",
            "Total: 54 x $500 = $27,000",
            "Market: CARFAC-RAAV $55-75/hr; 8hr day = $440-600/day. Your $500/day = $62.50/hr (within range).",
            "Responsibilities: Artistic research (8d), weekly dailies review (12d), prompts/constraints (6d), post-editing direction (20d), screening coordination (8d).",
        ]),
        ("LINE: Editor - $14,000", [
            "Scope: 3-5 short process films (3-8 min each), ~80-100 editing hours.",
            "Rate: $14,000 / ~90 hrs = ~$155/hr.",
            "Market: Toronto Arts Council media arts $150-200/hr; Vancouver indie $100-180/hr. Competitive.",
        ]),
        ("LINE: Systems/PM (Sector Support) - $60,000", [
            "Role: Systems lead + PM; ~20 hrs/week x 12 months (~1,040 hrs) at blended rate.",
            "Plus contracted UX/privacy advisory. Workflow design, documentation, multi-site coordination.",
        ]),
    ]
    row = 2
    for title, lines in blocks:
        ws.cell(row, 1, title).font = Font(bold=True)
        row += 1
        for line in lines:
            ws.cell(row, 1, line)
            row += 1
        row += 1
    ws.column_dimensions["A"].width = 90


def build_outcomes_metrics(ws):
    """Budget-to-outcomes mapping."""
    ws["A1"] = "BUDGET-TO-OUTCOMES MAPPING"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Budget Line", "Amount", "Outcome Metric", "Target", "Measurement"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    rows = [
        ("Lead artist fee", 27000, "Artistic process films", "3-5 films delivered", "Deliverables + filmmaker notes"),
        ("Participant honoraria", 18000, "Participants in co-creation", "25-30 artists retain authorship", "Registration + attendance + survey"),
        ("Editor fee", 14000, "Professional editing quality", "3-5 edited films", "Rubric (technical, creative, clarity)"),
        ("Accessibility", 2500, "Disabled/neurodivergent participation", "100+ community members", "Screening attendance + access feedback"),
        ("Systems/PM (SS)", 60000, "Toolkit completed & documented", "1 robust toolkit", "Facilitator modules, checklists, diagrams"),
        ("Sector workshops (venue etc)", 5000, "Disability-support orgs trained", "8-12 org staff trained", "Attendance + feedback + implementation tracking"),
    ]
    for i, row in enumerate(rows, 3):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
        ws.cell(i, 3, row[2])
        ws.cell(i, 4, row[3])
        ws.cell(i, 5, row[4])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 38


def build_rate_benchmarking(ws):
    """Comparison to industry standards."""
    ws["A1"] = "RATE BENCHMARKING - Artistic Creation"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Editor Fee: $14,000 for 3-5 short films, ~80-100 editing hours"
    ws["A3"] = "Your rate: $14,000 / ~90 hrs = ~$155/hour"
    ws["A5"] = "Industry (Vancouver 2024-25):"
    ws["A6"] = "Toronto Arts Council (media arts)"
    ws["B6"] = "$150-200/hr"
    ws["C6"] = "$155/hr OK"
    ws["A7"] = "Vancouver indie / freelance"
    ws["B7"] = "$100-180/hr"
    ws["C7"] = "Competitive"
    ws["A8"] = "Conclusion: Rate competitive and justified. References: CARFAC-RAAV 2024; TAC reports."
    ws.column_dimensions["A"].width = 45


def build_project_timeline(ws):
    """6-month timeline table (Artistic Creation)."""
    ws["A1"] = "ARTISTIC CREATION - 6-Month Timeline"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Month", "Week 1-2", "Week 3-4", "Deliverable", "Budget Spend"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    rows = [
        ("M1", "Onboarding, consent, setup", "Participant 1-2 sessions begin", "Kickoff done; 2 test videos", 10900),
        ("M2", "Participant 3-15 sessions", "Dailies review protocol running", "8 process videos draft", 8200),
        ("M3", "Participant 16-25 sessions", "Dailies analysis + creative notes", "Editing begins; 2 films in post", 16450),
        ("M4", "Participant 26-30 sessions", "Content review + accessibility planning", "2 edits ready for screening", 16450),
        ("M5", "Post-production finalization", "Community screening prep", "3-5 films finalized", 8200),
        ("M6", "Community screenings (2)", "Archiving + final reporting", "Screenings complete, archive done", 13800),
    ]
    for i, (m, w1, w2, deliv, spend) in enumerate(rows, 3):
        ws.cell(i, 1, m)
        ws.cell(i, 2, w1)
        ws.cell(i, 3, w2)
        ws.cell(i, 4, deliv)
        ws.cell(i, 5, spend)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 12


def build_revision_log(ws):
    """Budget revision / audit trail."""
    ws["A1"] = "BUDGET REVISION LOG"
    ws["A1"].font = Font(bold=True, size=12)
    headers = ["Date", "Sheet", "Line Item", "Change", "Rationale"]
    for col, h in enumerate(headers, 1):
        ws.cell(2, col, h).font = Font(bold=True)
    rows = [
        ("2026-03-03", "AC_Expenses", "Equipment purchase", "Added: $2,500", "SSDs, card readers for 6-month capture"),
        ("2026-03-04", "AC_Expenses", "Contingency", "8.7% ($6,500)", "Realistic planning for device replacement, etc."),
        ("2026-03-05", "SS_Expenses", "Systems/PM fee", "Enhanced note (no $ change)", "Added hourly breakdown for transparency"),
        ("2026-03-05", "Cover", "Pre-submission note", "Added advice", "Contact program officer re $100K ceiling"),
    ]
    for i, row in enumerate(rows, 3):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 45


def build_assumptions_constraints(ws):
    """Key assumptions and constraints."""
    ws["A1"] = "KEY ASSUMPTIONS & CONSTRAINTS"
    ws["A1"].font = Font(bold=True, size=12)
    items = [
        ("25-30 participants complete full 6-month engagement.", "If <20 enroll, editing scope reduced (3 films vs 5).", "Early onboarding, clear expectations."),
        ("Lead artist available 2 days/week for 6 months.", "If less available, editing direction reduced.", "Contract secured before funding approval."),
        ("Partner sites provide free/low-cost space access.", "If costs spike, travel budget reallocated.", "Letter of commitment from partners required."),
        ("Accessibility services at budgeted rates.", "If interpreters/captioners cost more, contingency deployed.", "Early vendor quotes obtained."),
    ]
    ws["A2"] = "Assumption"
    ws["B2"] = "Constraint"
    ws["C2"] = "Mitigation"
    for c in range(1, 4):
        ws.cell(2, c).font = Font(bold=True)
    for i, row in enumerate(items, 3):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
        ws.cell(i, 3, row[2])
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 38


def build_supporting_documents(ws):
    """Checklist of supporting documents."""
    ws["A1"] = "SUPPORTING DOCUMENTS CHECKLIST"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "With budget submission:"
    ws["A2"].font = Font(bold=True)
    docs = [
        "Letter of support from [Partner Org A]",
        "Letter of support from [Partner Org B]",
        "Quote from accessibility services (captioning, interpretation)",
        "Quote from post-production editor (confirming $14,000 scope)",
        "Proof of business registration (Punkeye Pictures)",
        "CARFAC-RAAV rate justification (for peer review)",
    ]
    for i, d in enumerate(docs, 3):
        ws.cell(i, 1, "[ ] " + d)
    ws["A10"] = "Before funds disbursed:"
    ws["A10"].font = Font(bold=True)
    ws.cell(11, 1, "[ ] Signed participant consent forms")
    ws.cell(12, 1, "[ ] Signed contractor agreements (lead artist, editor)")
    ws.cell(13, 1, "[ ] Insurance certificate")
    ws["A14"] = "After project completion:"
    ws["A14"].font = Font(bold=True)
    ws.cell(15, 1, "[ ] Receipts for all expenditures")
    ws.cell(16, 1, "[ ] Final report with outcomes data")
    ws.cell(17, 1, "[ ] Participant feedback forms (anonymized)")
    ws.column_dimensions["A"].width = 58


def main():
    notes = ensure_notes_under_250(notes_under_250())
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove default sheet and create named sheets
    wb.remove(wb.active)
    wb.create_sheet("Cover", 0)
    wb.create_sheet("AC_Expenses", 1)
    wb.create_sheet("AC_Income", 2)
    wb.create_sheet("SS_Expenses", 3)
    wb.create_sheet("SS_Income", 4)
    wb.create_sheet("Budget_Checks", 5)
    wb.create_sheet("AC_CashFlow", 6)
    wb.create_sheet("SS_CashFlow", 7)
    wb.create_sheet("Risk_Analysis", 8)
    wb.create_sheet("InKind_Mapping", 9)
    wb.create_sheet("Expense_Justification", 10)
    wb.create_sheet("Outcomes_Metrics", 11)
    wb.create_sheet("Rate_Benchmarking", 12)
    wb.create_sheet("Project_Timeline", 13)
    wb.create_sheet("Revision_Log", 14)
    wb.create_sheet("Assumptions_Constraints", 15)
    wb.create_sheet("Supporting_Documents", 16)
    build_cover(wb["Cover"])
    build_ac_expenses(wb["AC_Expenses"], notes)
    build_ac_income(wb["AC_Income"], notes)
    build_ss_expenses(wb["SS_Expenses"], notes)
    build_ss_income(wb["SS_Income"], notes)
    build_budget_checks(wb["Budget_Checks"])
    build_ac_cashflow(wb["AC_CashFlow"])
    build_ss_cashflow(wb["SS_CashFlow"])
    build_risk_analysis(wb["Risk_Analysis"])
    build_inkind_mapping(wb["InKind_Mapping"])
    build_expense_justification(wb["Expense_Justification"])
    build_outcomes_metrics(wb["Outcomes_Metrics"])
    build_rate_benchmarking(wb["Rate_Benchmarking"])
    build_project_timeline(wb["Project_Timeline"])
    build_revision_log(wb["Revision_Log"])
    build_assumptions_constraints(wb["Assumptions_Constraints"])
    build_supporting_documents(wb["Supporting_Documents"])
    add_data_validation(wb)
    add_conditional_formatting(wb)
    # Force Excel to recalculate all formulas when the file is opened (fixes formulas showing blank in some viewers)
    wb.calculation.fullCalcOnLoad = True
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    # Verify formulas were written
    _verify_formulas(OUTPUT_FILE)


def _verify_formulas(path):
    """Re-open the saved file and verify key cells contain formulas (data_type 'f')."""
    try:
        wb = load_workbook(path, data_only=False)
        checks = [
            ("AC_Expenses", "C9", "=SUM(C2:C8)"),
            ("AC_Expenses", "C21", "=SUM(C11:C20)"),
            ("AC_Expenses", "C25", "=SUM(C23:C24)"),
            ("AC_Expenses", "C32", "=SUM(C9,C21,C25,C26,C27,C30)"),
            ("AC_Income", "B9", "=SUM(B2:B8)"),
            ("AC_Income", "B10", "=B9-"),
            ("SS_Expenses", "C9", "=SUM(C5:C8)"),
            ("SS_Expenses", "C25", "=SUM(C3,C9,C13,C16,C20,C23)"),
            ("SS_Income", "B9", "=SUM(B2:B8)"),
            ("Budget_Checks", "B2", "=IF("),
        ]
        missing = []
        for sheet_name, cell_ref, expected_start in checks:
            ws = wb[sheet_name]
            cell = ws[cell_ref]
            val = cell.value
            if val is None or (isinstance(val, str) and not val.strip().startswith("=")):
                missing.append(f"{sheet_name}!{cell_ref}")
            elif isinstance(val, str) and expected_start and not val.lstrip().startswith(expected_start):
                if expected_start != "=IF(" or "IF(" not in val:
                    missing.append(f"{sheet_name}!{cell_ref} (unexpected: {str(val)[:50]}...)")
        wb.close()
        if missing:
            print("WARNING: These cells may not have formulas:", missing)
        else:
            print("OK: Key formula cells verified.")
    except Exception as e:
        print(f"Note: Could not verify formulas ({e}). Open the file in Excel and press Ctrl+Alt+F9 to recalculate.")


if __name__ == "__main__":
    main()
